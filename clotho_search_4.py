# 4 再次进行fine-grained embedding
import logging
import torch
import data
import torch.nn as nn
from models import imagebind_model
from models.imagebind_model import ModalityType, load_module
from models import lora as LoRA
import pandas as pd
from torchvision import transforms
from torchvision.datasets import ImageNet
from torch.utils.data import DataLoader
import numpy as np
import matplotlib.pyplot as plt
from torch.nn.parallel import DataParallel
from api.clotho_text2audio import ClothoTextDataset
from api.clotho import ClothoDataset
logging.basicConfig(level=logging.INFO, force=True)
import os
csv_file_path = "/home/u2021010261/data/cdq/clotho/clotho_captions_evaluation.csv"
data_dir="/home/u2021010261/data/cdq/clotho/evaluation/"
f_s=os.listdir(data_dir)
print(len(f_s))
pf=pd.read_csv(csv_file_path,sep=',') # 假设数据集以CSV文件形式提供
text_list = pf[['caption_1', 'caption_2', 'caption_3', 'caption_4', 'caption_5']].values.flatten().tolist()
audio_list=pf[['file_name']].values.flatten().tolist()
audio_path=[data_dir+file for file in audio_list]
device = "cuda:0" if torch.cuda.is_available() else "cpu"
device_ids = [0,1,2] 
Clotho_dataset = ClothoTextDataset(csv_file=csv_file_path,device=device)
batch_size=128
test_dl = DataLoader(dataset=Clotho_dataset, batch_size=batch_size, shuffle=False, drop_last=False,
        num_workers=4, pin_memory=True, persistent_workers=True)
test_dl2 = DataLoader(dataset=Clotho_dataset, batch_size=1, shuffle=False, drop_last=False,
             num_workers=4, pin_memory=True, persistent_workers=True)
batch_size=1        
K=30
counts_r1=np.array([])
counts_r10=np.array([])
topk1=[1,5, 10, 20, 30, 40, 50, 60,70,80,90,100,110,120,130,300,400,500,600,700,800,900,1000]
counts_rs = {}
for k in topk1:
    if k<=K:
        counts_rs[f'counts_r{k}'] = np.array([])
embeddings={}
embeddings_12={}
import pickle


# 文件路径
file_path = "shortlist_data.pkl"
# 提取数据
with open(file_path, 'rb') as f:
    shortlist = pickle.load(f)
    shortlist_item = pickle.load(f)
fine_model = imagebind_model.imagebind_huge(pretrained=True)
fine_model=fine_model.to(device)
with torch.no_grad():
        checkpoint = torch.load(f'parameters/audio/trunks+post/embeddings_{12}.pth')
        # 获取模型参数和张量
        embeddings_12[ModalityType.AUDIO]= checkpoint['audio_embeddings']
        print(1)

for batch_idx, (x, target) in enumerate(test_dl2):
            embeddings_AUDIO={}
            target = target.to(device)
            inputs = {
                ModalityType.TEXT: data.load_and_transform_text(x, device)
                #ModalityType.AUDIO: data.load_and_transform_audio_data(shortlist[f'counts_r{K}'][batch_idx*batch_size],device=device)
            }
            for item in shortlist_item[f'counts_r{K}'][batch_idx*batch_size]:
                if embeddings_AUDIO:
                    embeddings_AUDIO[ModalityType.AUDIO]=torch.cat([embeddings_AUDIO[ModalityType.AUDIO], embeddings_12['audio'][item].unsqueeze(0).to(embeddings_AUDIO[ModalityType.AUDIO].device)], dim=0)        
                else:
                    embeddings_AUDIO[ModalityType.AUDIO] = embeddings_12['audio'][item].unsqueeze(0)
            embeddings = fine_model(inputs)
            match_value_1 = embeddings[ModalityType.TEXT] @ embeddings_AUDIO[ModalityType.AUDIO].to(device).T 
            result_1 = torch.softmax(match_value_1, dim=-1)
            _, predicted = torch.max(result_1, dim=-1)
            top_indices_list = [torch.topk(result_1, k=k, dim=-1)[1] if k <= K else None for k in topk1]
            r1=0
            r10=0
            predicted=torch.Tensor([shortlist_item[f'counts_r{K}'][batch_idx*batch_size][predicted]])
        
            #_, topk_indices = torch.topk(result_1, k=10, dim=-1)
            #top_indices_list = [torch.topk(result_1, k=k, dim=-1)[1] if k <= len(shortlist_item[f'counts_r{K}'][batch_idx]) else None for k in topk1]
            for k, top_indices, counts_rk in zip(topk1, top_indices_list, counts_rs):
                if k == 1:
                        counts_rs[counts_rk] = np.concatenate([counts_rs[counts_rk], [int(predicted[i] == target.to(predicted.device)) for i in range(len(predicted))]])
                elif k<=K:            
                        results_k=[]
                        for i_k in range(k):
                            results_k.append(shortlist_item[f'counts_r{K}'][batch_idx][top_indices[0][i_k].item()])
                        counts_rs[counts_rk] = np.concatenate([counts_rs[counts_rk], [int(any(results_k[i] == target.to(predicted.device))) for i in range(len(results_k))]])
                        
                else :
                    break
            r1=(np.sum(counts_rs['counts_r1']))/len(counts_rs['counts_r1'])
            r10=(np.sum(counts_rs['counts_r10']))/len(counts_rs['counts_r1']) 
        
            logging.info(f"batch_idx = {batch_idx}, r1={r1},r10={r10}, test_total = {len(counts_r1)}")
    
results=[]
lists=[]
for counts in counts_rs:
    correct=np.sum(counts_rs[counts] == 1)/len(counts_rs[counts])
    results.append(str(correct))
    lists.append(str(counts))
# a_block=len(model_1.modality_trunks["audio"].blocks)


import openpyxl
from openpyxl import Workbook
# 数据
data1 = [
    lists,
    results
]

# 打开Excel文件
#workbook = openpyxl.load_workbook('topk-clotho-search.xlsx')
workbook=Workbook()
# 选择或创建工作表
sheet_name = 'Sheet0'  # 请根据实际情况修改工作表名称
sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

# 确定插入的起始行
start_row = sheet.max_row + 1

# 将数据插入Excel
for row in data1:
    sheet.append(row)

# 保存修改后的Excel文件
workbook.save('end_to_end.xlsx')