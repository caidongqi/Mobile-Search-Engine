import logging
import torch
import json
import data
import torchvision
import torchmetrics
import torch.nn as nn
from models import imagebind_model
from models.imagebind_model import ModalityType, load_module
from models import lora as LoRA
import pandas as pd
from torchvision import transforms
from torchvision.datasets import ImageNet
from torch.utils.data import DataLoader
import numpy as np
import matplotlib.pyplot as plt
from torch.nn.parallel import DataParallel
from api.clotho_text2audio import ClothoTextDataset
from api.clotho import ClothoDataset
logging.basicConfig(level=logging.INFO, force=True)
import os
csv_file_path = "/data/air/pc/Mobile-Search-Engine/datasets/clotho/clotho_captions_evaluation.csv"
data_dir="/data/air/pc/Mobile-Search-Engine/datasets/clotho/evaluation"
f_s=os.listdir(data_dir)
print(len(f_s))
pf=pd.read_csv(csv_file_path,sep=',') # 假设数据集以CSV文件形式提供
text_list = pf[['caption_1', 'caption_2', 'caption_3', 'caption_4', 'caption_5']].values.flatten().tolist()
audio_list=pf[['file_name']].values.flatten().tolist()
audio_path=["/data/air/pc/Mobile-Search-Engine/datasets/clotho/evaluation/"+file for file in audio_list]
embeddings={}


# # 1.存储n层 n=1
# N=1
device = "cuda:4" if torch.cuda.is_available() else "cpu"
device_ids = [0,1,3,4,6,7] 
Clotho_dataset = ClothoTextDataset(csv_file=csv_file_path,device=device)
# batch_size=128
# test_dl = DataLoader(dataset=Clotho_dataset, batch_size=batch_size, shuffle=False, drop_last=False,
#         num_workers=4, pin_memory=True, persistent_workers=True)

# with torch.no_grad():
#         checkpoint = torch.load(f'parameters/audio/trunks+post/embeddings_{N}.pth')
#         # 获取模型参数和张量
#         embeddings[ModalityType.AUDIO]= checkpoint['audio_embeddings']
#         print(1)





# # 2.每个数据动态存储 m 层
# import torch.nn.functional as F

# # 定义模型结构
# class MyModel(nn.Module):
#     def __init__(self, input_size, output_size):
#         super(MyModel, self).__init__()
#         self.fc1 = nn.Linear(input_size, 256)
#         self.fc2 = nn.Linear(256, 64)
#         self.fc3 = nn.Linear(64, output_size)

#     def forward(self, x):
#         x = F.relu(self.fc1(x))
#         x = F.relu(self.fc2(x))
#         x = self.fc3(x)
#         return x

# # 创建模型实例
# input_size = 1024  # 根据embedding的大小确定输入层大小
# output_size = 12  # 根据层数的范围确定输出层大小
# predict_model = MyModel(input_size, output_size)  # 请确保input_size和output_size已定义
# predict_model.to(device)
# # 加载已保存的模型参数
# predict_model.load_state_dict(torch.load('model_trunks1&2_parameters.pth'))
# models = []
# import math
# for n in range(0, output_size ):
#     model = imagebind_model.imagebind_huge(pretrained=True, audio_num_blocks=n)
#     device1 = "cuda:0" if torch.cuda.is_available() else "cpu"
#     model.to(device_ids[math.ceil((n-1)/2)])
#     models.append(model)
# layers=[]
# for embedding_item in embeddings[ModalityType.AUDIO]:
#         embedding_item=embedding_item.to(device)
#         layer=predict_model(embedding_item.float())
#         _, layer1 = torch.max(layer, 0)
#         layers.append(layer1)
# import pickle

# # 定义文件路径
# file_path = "layers.pkl"

# # 保存数据到文件
# with open(file_path, 'wb') as f:
#     pickle.dump(layers, f)
# embedding_dynamic={}

# with torch.no_grad():
#   for i in range(len(embeddings[ModalityType.AUDIO])):
        
#         inputs = {
#         ModalityType.AUDIO: data.load_and_transform_audio_data2(audio_path[i],device=device_ids[math.ceil((layers[i]-1)/2)])
#         }
        
#         current_embeddings = models[layers[i]](inputs)[ModalityType.AUDIO]

#         if embedding_dynamic:
#                 embedding_dynamic[ModalityType.AUDIO] = torch.cat([embedding_dynamic[ModalityType.AUDIO], current_embeddings.to(embedding_dynamic[ModalityType.AUDIO].device)], dim=0)
#         else:
#                 embedding_dynamic[ModalityType.AUDIO] = current_embeddings

#         del current_embeddings
# torch.save({
#         'audio_embeddings': embedding_dynamic[ModalityType.VISION]
#     }, f'parameters/dynamic/image/embeddings_{N}.pth')

# embedding_dynamic={}
# with torch.no_grad():
#         checkpoint = torch.load(f'parameters/dynamic/image/embeddings_{N}.pth')
#         # 获取模型参数和张量
#         embedding_dynamic[ModalityType.AUDIO]= checkpoint['audio_embeddings']
#         print(1)








# #3 根据query进行match到前k个数据
fine_model = imagebind_model.imagebind_huge(pretrained=True)
fine_model=fine_model.to(device)
# top_k={}
# topk1=[1,5, 10, 20, 30, 40, 50, 60,70,80,90,100,110,120,130,300,400,500,600,700,800,900,1000]
# counts_rs = {}
# shortlist={}
# shortlist_item={}
# for k in topk1:
#         counts_rs[f'counts_r{k}'] = np.array([])
#         shortlist[f'counts_r{k}']=[]
#         shortlist_item[f'counts_r{k}']=[]
    
# K=10
# counts_r1=np.array([])
# counts_r10=np.array([])
# for batch_idx, (x, target) in enumerate(test_dl):
#         target = target.to(device)
#         inputs = {
#         ModalityType.TEXT: data.load_and_transform_text(x, device)
#         }
#         fine_embeddings= fine_model(inputs)[ModalityType.TEXT].to(embedding_dynamic[ModalityType.AUDIO].device)
#         match_value_1 = fine_embeddings @ embedding_dynamic[ModalityType.AUDIO].T 
#         result_1 = torch.softmax(match_value_1, dim=-1)
#         _, predicted = torch.max(result_1, dim=-1)
#         _, topk_indices = torch.topk(result_1, k=K, dim=-1)
#         counts_r1 = np.concatenate([counts_r1, [int(predicted[i] == target[i].to(predicted.device)) for i in range(len(predicted))]])
#             # #counts_r1 = np.concatenate([counts_r1, [any(predicted[i] == target[i]) for i in range(len(predicted))]])
#             # #topk_indices=topk_indices.T
#         counts_r10=np.concatenate([counts_r10, [int(any(topk_indices[i] == target[i].to(predicted.device))) for i in range(len(target))]])
        
#         top_indices_list = [torch.topk(result_1, k=k, dim=-1)[1] for k in topk1]
        
#         for k, top_indices, counts_r in zip(topk1, top_indices_list, counts_rs):
#                 if k == 1:
#                         counts_rs[counts_r] = np.concatenate([counts_rs[counts_r], [int(predicted[i] == target[i].to(predicted.device)) for i in range(len(predicted))]])
                        
#                 else:
#                         counts_rs[counts_r] = np.concatenate([counts_rs[counts_r], [int(any(top_indices[i] == target[i].to(predicted.device))) for i in range(len(target))]])
#                         # for i,row in enumerate(top_indices):
#                         #     list=[]
#                         #     list_item=[]
#                         #     for item in row:
#                         #         list.append(audio_path[item])
#                         #         list_item.append(item.item())
#                         #     shortlist[counts_r].append(list)
#                         #     shortlist_item[counts_r].append(list_item)

#                 for i,row in enumerate(top_indices):
#                     list=[]
#                     list_item=[]
#                     for item in row:
#                         list.append(audio_path[item])
#                         list_item.append(item.item())
#                     shortlist[counts_r].append(list)
#                     shortlist_item[counts_r].append(list_item)
            
# import pickle

# # 假设你有一个文件路径用于保存数据
# file_path = "shortlist_data.pkl"

# # 保存 shortlist 和 shortlist_item 到本地文件
# with open(file_path, 'wb') as f:
#     pickle.dump(shortlist, f)
#     pickle.dump(shortlist_item, f)

# print("Data saved successfully to", file_path)
import pickle

# 文件路径
file_path = "shortlist_data.pkl"

# 提取数据
with open(file_path, 'rb') as f:
    shortlist = pickle.load(f)
    shortlist_item = pickle.load(f)

# # 打印结果
# print("Shortlist:", shortlist)
# print("Shortlist Item:", shortlist_item)





# 4 再次进行fine-grained embedding
test_dl2 = DataLoader(dataset=Clotho_dataset, batch_size=1, shuffle=False, drop_last=False,
             num_workers=4, pin_memory=True, persistent_workers=True)
batch_size=1        
K=30
counts_r1=np.array([])
counts_r10=np.array([])
topk1=[1,5, 10, 20, 30, 40, 50, 60,70,80,90,100,110,120,130,300,400,500,600,700,800,900,1000]
counts_rs = {}
for k in topk1:
    if k<=K:
        counts_rs[f'counts_r{k}'] = np.array([])
embeddings={}
embeddings_12={}

with torch.no_grad():
        checkpoint = torch.load(f'parameters/audio/trunks+post/embeddings_{12}.pth')
        # 获取模型参数和张量
        embeddings_12[ModalityType.AUDIO]= checkpoint['audio_embeddings']
        print(1)

for batch_idx, (x, target) in enumerate(test_dl2):
            embeddings_AUDIO={}
            target = target.to(device)
            inputs = {
                ModalityType.TEXT: data.load_and_transform_text(x, device)
                #ModalityType.AUDIO: data.load_and_transform_audio_data(shortlist[f'counts_r{K}'][batch_idx*batch_size],device=device)
            }
            for item in shortlist_item[f'counts_r{K}'][batch_idx*batch_size]:
                if embeddings_AUDIO:
                    embeddings_AUDIO[ModalityType.AUDIO]=torch.cat([embeddings_AUDIO[ModalityType.AUDIO], embeddings_12['audio'][item].unsqueeze(0).to(embeddings_AUDIO[ModalityType.AUDIO].device)], dim=0)        
                else:
                    embeddings_AUDIO[ModalityType.AUDIO] = embeddings_12['audio'][item].unsqueeze(0)
            embeddings = fine_model(inputs)
            match_value_1 = embeddings[ModalityType.TEXT] @ embeddings_AUDIO[ModalityType.AUDIO].to(device).T 
            result_1 = torch.softmax(match_value_1, dim=-1)
            _, predicted = torch.max(result_1, dim=-1)
            top_indices_list = [torch.topk(result_1, k=k, dim=-1)[1] if k <= K else None for k in topk1]
            r1=0
            r10=0
            predicted=torch.Tensor([shortlist_item[f'counts_r{K}'][batch_idx*batch_size][predicted]])
        
            #_, topk_indices = torch.topk(result_1, k=10, dim=-1)
            #top_indices_list = [torch.topk(result_1, k=k, dim=-1)[1] if k <= len(shortlist_item[f'counts_r{K}'][batch_idx]) else None for k in topk1]
            for k, top_indices, counts_rk in zip(topk1, top_indices_list, counts_rs):
                if k == 1:
                        counts_rs[counts_rk] = np.concatenate([counts_rs[counts_rk], [int(predicted[i] == target.to(predicted.device)) for i in range(len(predicted))]])
                elif k<=K:            
                        results_k=[]
                        for i_k in range(k):
                            results_k.append(shortlist_item[f'counts_r{K}'][batch_idx][top_indices[0][i_k].item()])
                        counts_rs[counts_rk] = np.concatenate([counts_rs[counts_rk], [int(any(results_k[i] == target.to(predicted.device))) for i in range(len(results_k))]])
                        
                else :
                    break
            r1=(np.sum(counts_rs['counts_r1']))/len(counts_rs['counts_r1'])
            r10=(np.sum(counts_rs['counts_r10']))/len(counts_rs['counts_r1']) 
        
            logging.info(f"batch_idx = {batch_idx}, r1={r1},r10={r10}, test_total = {len(counts_r1)}")
    
results=[]
lists=[]
for counts in counts_rs:
    correct=np.sum(counts_rs[counts] == 1)/len(counts_rs[counts])
    results.append(str(correct))
    lists.append(str(counts))
# a_block=len(model_1.modality_trunks["audio"].blocks)


import openpyxl
from openpyxl import Workbook
# 数据
data1 = [
    lists,
    results
]

# 打开Excel文件
#workbook = openpyxl.load_workbook('topk-clotho-search.xlsx')
workbook=Workbook()
# 选择或创建工作表
sheet_name = 'Sheet0'  # 请根据实际情况修改工作表名称
sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

# 确定插入的起始行
start_row = sheet.max_row + 1

# 将数据插入Excel
for row in data1:
    sheet.append(row)

# 保存修改后的Excel文件
workbook.save('end_to_end.xlsx')