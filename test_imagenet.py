import logging
import torch
import data
import torchvision
import torchmetrics
import time
import csv

from models import imagebind_model
from models import lumen_model
from models import lumen6_model_test
from models.imagebind_model import ModalityType, load_module
from models.lumen6_model_test import ModalityType, load_module
from models.lumen_model import ModalityType, load_module
from models import lora_lumen as LoRA

from pycocotools.coco import COCO
from torchvision import transforms
from torch.utils.data import DataLoader
from torchvision.datasets import CocoDetection

from torchvision import transforms
from torchvision.datasets import ImageNet
from torch.utils.data import DataLoader
from api.imagenet import ImageNetDataset
from api.coco import CoCoDataset
from metrics.accuracy import Accuracy
import openpyxl
logging.basicConfig(level=logging.INFO, force=True)
import numpy as np
print("test_imagebet")
lora = True
linear_probing = False
#device="cpu"
#device = "cuda:6" if torch.cuda.is_available() else "cpu"
load_head_post_proc_finetuned = False
#imagenet_datadir = "/data/yx/ImageBind/.datasets/imagenet"
imagenet_datadir = "/data/air/pc/Mobile-Search-Engine/.datasets/one_imagenet"
coco_datadir="/data/air/pc/ImageBind/dataset/val2017/val2017"
test_coco="/data/air/pc/ImageBind/dataset/tempoimage"
datadir1='/data/air/pc/Mobile-Search-Engine/datasets/imagenet-10'
datadir2="/data/air/pc/i-Code/i-Code-V3/dataset/tempo-10"
coco_annotation_file='/data/air/pc/ImageBind/dataset/annotations_trainval2017/annotations/instances_val2017.json'
lora_dir = '/data/air/pc/Mobile-Search-Engine/.checkpoints/lora/exp_6'

import argparse

# 创建解析器
parser = argparse.ArgumentParser(description="Your script description")

# 添加命令行参数
#parser.add_argument("audio_num_blocks", type=int, help="Number of audio blocks")

# parser.add_argument("--audio_num_blocks", default=12, type=int, help="Number of audio blocks")
parser.add_argument("--device", type=str, default="cuda:5", help="Device to use (cuda:2 or cpu)")
parser.add_argument("--vision_num_blocks", default=32,type=int, help="Number of audio blocks")
# 解析命令行参数
args = parser.parse_args()

# 获取 audio_num_blocks 的值
vision_num_blocks=args.vision_num_blocks
device = args.device



assert not (linear_probing and lora), \
            "Linear probing is a subset of LoRA training procedure for ImageBind. " \
            "Cannot set both linear_probing=True and lora=True. "

if lora and not load_head_post_proc_finetuned:
    # Hack: adjust lora_factor to the `max batch size used during training / temperature` to compensate missing norm
    lora_factor = 4 / 0.07
else:
    # This assumes proper loading of all params but results in shift from original dist in case of LoRA
    lora_factor = 1
# Instantiate model
model=lumen6_model_test.imagebind_huge(pretrained=True,vision_num_blocks_1=31,vision_num_blocks_2=1)
#model = imagebind_model.imagebind_huge(pretrained=True,vision_num_blocks=vision_num_blocks)


if lora:
    model.modality_trunks_1.update(LoRA.apply_lora_modality_trunks(model.modality_trunks_1, rank=4,
                                                                              layer_idxs=None,
                                                                              modality_names=[ModalityType.TEXT, ModalityType.VISION]))
    # Load LoRA params if found
    # #LoRA.load_lora_modality_trunks(model.modality_trunks_1,
    #                                checkpoint_dir=lora_dir)
    LoRA.load_lora_modality_trunks(model.modality_trunks_1, checkpoint_dir=lora_dir, postfix = "_trunk1_last")
           
    model.modality_trunks_2.update(LoRA.apply_lora_modality_trunks(model.modality_trunks_2, rank=4,
                                                                              layer_idxs=None,
                                                                              modality_names=[ModalityType.TEXT, ModalityType.VISION]))
    # Load LoRA params if found
    LoRA.load_lora_modality_trunks(model.modality_trunks_2, checkpoint_dir=lora_dir, postfix = "_trunk2_last")
       
    # model.modality_trunks.update(
    #     LoRA.apply_lora_modality_trunks(model.modality_trunks, rank=4,
    #                                     layer_idxs={ModalityType.TEXT: [ 1, 2, 3, 4, 5, 6, 7, 8],
    #                                                 ModalityType.VISION: [1, 2, 3, 4, 5, 6, 7, 8]},
    #                                     modality_names=[ModalityType.TEXT, ModalityType.VISION]))

    # # Load LoRA params if found
    # LoRA.load_lora_modality_trunks(model.modality_trunks,
    #                                checkpoint_dir=lora_dir)

    if load_head_post_proc_finetuned:
        # Load postprocessors & heads
        load_module(model.modality_postprocessors, module_name="postprocessors",
                    checkpoint_dir=lora_dir)
        load_module(model.modality_heads, module_name="heads",
                    checkpoint_dir=lora_dir)
elif linear_probing:
    # Load heads
    load_module(model.modality_heads, module_name="heads",
                checkpoint_dir=lora_dir)

model.eval()
model.to(device)


def run_inference():
    data_transform = transforms.Compose(
        [
            transforms.Resize(
                224, interpolation=transforms.InterpolationMode.BICUBIC
            ),
            transforms.CenterCrop(224),
            transforms.ToTensor(),
            transforms.Normalize(
                mean=(0.48145466, 0.4578275, 0.40821073),
                std=(0.26862954, 0.26130258, 0.27577711),
            ),
        ]
    )

    #test_ds1=CoCoDataset(datadir=coco_datadir, annFile=coco_annotation_file,transform=data_transform)
    test_ds = ImageNetDataset(datadir=imagenet_datadir, split="val", transform=data_transform)
    #test_dl1=DataLoader(dataset=test_ds1, batch_size=64, shuffle=False, drop_last=False, num_workers=4, pin_memory=True, persistent_workers=True)
    test_dl = DataLoader(dataset=test_ds, batch_size=1, shuffle=False, drop_last=False, num_workers=4, pin_memory=True, persistent_workers=True)
    import os
    test_acc = Accuracy(task="multiclass", num_classes=1000, average="micro",device=device)
    file_count = sum(1 for entry in os.scandir(coco_datadir) if entry.is_file())
    test_correct = 0
    test_total = 0
    import data
    total_correct=0
    print(len(test_ds))

   
  
    topk1=[1,5, 10, 20, 30, 40, 50, 60,70,80,90,100,110,120,130,300,400,500,600]
    counts_rs = {}
    for k in topk1:
                counts_rs[f'counts_r{k}'] = np.array([])
    with torch.no_grad():
        for batch_idx, (x, target,imgs) in enumerate(test_dl):
            
            x = x.to(device)
            target=[t.to(device) for t in target]
            print(imgs)
            inputs = {
                ModalityType.VISION: x,
                ModalityType.TEXT: data.load_and_transform_text(test_ds.text_list, device),
            }
            embeddings = model(inputs)
            match_value_1 = embeddings[ModalityType.VISION]@embeddings[ModalityType.TEXT].T * (lora_factor if lora else 1)
            print(match_value_1.shape)
            # num=5
            # topk_indices = torch.topk(match_value_1, num, dim=-1).indices
            # # 获取每行最大值的索引
            # max_indices = torch.argmax(match_value_1, dim=1)
            result_1 = torch.softmax(match_value_1, dim=-1)
            _, predicted = torch.max(result_1, -1)
            #topk=[5, 10, 20, 30, 40, 50, 60,70,80,90,100,110,120,130,300,400,500,600]
            
            top_indices_list = [torch.topk(result_1, k=k, dim=-1)[1] for k in topk1]
            
            for k, top_indices, counts_r in zip(topk1, top_indices_list, counts_rs):
                if k == 1:
                    counts_rs[counts_r] = np.concatenate([counts_rs[counts_r], [int(predicted[i] == target[i].to(predicted.device)) for i in range(len(predicted))]])
                else:
                    counts_rs[counts_r] = np.concatenate([counts_rs[counts_r], [int(any(top_indices[i] == target[i].to(predicted.device))) for i in range(len(target))]])

            logging.info(f"batch_idx = {batch_idx}, test_correct = {np.sum(counts_rs['counts_r1'] == 1)/len(counts_rs['counts_r1'])}, test_total = {np.sum(counts_rs['counts_r5'] == 1)/len(counts_rs['counts_r1'])}, Accuracy = {np.sum(counts_rs['counts_r10'] == 1)/len(counts_rs['counts_r1'])}")
            # if(np.sum(counts_rs['counts_r1'] == 1)):
            #     print(f"batch_idx不为0:{imgs}")
            #     embed=embeddings[ModalityType.VISION]
            #     print(embed)
            #     torch.save(embed, "output_tensor.pt")
            #     return imgs
            # for i in range(top_k_indices.size(0)):
            #     row_tensor1 = top_k_indices[i, :]
            #     row_tensor1= torch.add(row_tensor1, 1)
            #     # 提取每个子tensor的第i列
            #     column_values = [tensor.item() for tensor in target]
            #     extracted_values = [str(item) for item in column_values]

            #     # 组成新的tensor
            #     new_tensor = torch.cat([torch.tensor(extracted_values)])
            #     new_tensor=torch.unique(new_tensor.to(device))
            #     intersection = torch.unique(torch.cat((row_tensor1, new_tensor)))
            #     #intersection = torch.intersection(row_tensor1, new_tensor)
            #     if len(intersection)<len(row_tensor1) + len(new_tensor):
            #             correct+=1
           
    #total_acc = test_acc.compute()
    
    indices = [i for i, value in enumerate(counts_rs['counts_r1'] ) if value != 0]
    print(indices)
    results=[]
    lists=[]
    for counts in counts_rs:
        correct=np.sum(counts_rs[counts] == 1)/len(counts_rs[counts])
        results.append(str(correct))
        lists.append(counts)
    
    #将时间记录下来
    # v_block=len(model.modality_trunks["vision"].blocks)
    # t_block=len(model.modality_trunks["text"].blocks)
    # a_block=len(model.modality_trunks["audio"].blocks)
    # i_block=len(model.modality_trunks["imu"].blocks)

    # import openpyxl
    # r1=[str(v_block),str(t_block),str(a_block),str(i_block)]
    # list1=['vision层数', 'text层数', 'audio层数', 'imu层数']
    # list=list1+lists
    # results=r1+results
    # # 数据
    # data = [
    #      list,results
    # ]
    
    # # 打开Excel文件
    # workbook = openpyxl.load_workbook('topk.xlsx')

    # # 选择或创建工作表
    # sheet_name = 'Sheet1'  # 请根据实际情况修改工作表名称
    # sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

    # # 确定插入的起始行
    # start_row = sheet.max_row + 1

    # # 将数据插入Excel
    # for row in data:
    #     sheet.append(row)

    # # 保存修改后的Excel文件
    # workbook.save('topk.xlsx')

    
    # # 写入CSV文件
    # with open(filename, 'w', newline='') as file:
    #     writer = csv.writer(file)
    #     writer.writerows(data)

    # print(f"数据已写入CSV文件: {filename}")

   
    
    


if __name__ == "__main__":
    run_inference()
