import logging
import torch
import json
import data
import torchvision
import torchmetrics
import torch.nn as nn
from models import imagebind_model
from models.imagebind_model import ModalityType, load_module
from models import lora as LoRA
import pandas as pd
from torchvision import transforms
from torchvision.datasets import ImageNet
from torch.utils.data import DataLoader
import numpy as np
import matplotlib.pyplot as plt
from torch.nn.parallel import DataParallel
from api.clotho_text2audio import ClothoTextDataset
from api.clotho import ClothoDataset
logging.basicConfig(level=logging.INFO, force=True)
import os
csv_file_path = "/data/air/pc/Mobile-Search-Engine/datasets/clotho/clotho_captions_evaluation.csv"
data_dir="/data/air/pc/Mobile-Search-Engine/datasets/clotho/evaluation"
f_s=os.listdir(data_dir)
print(len(f_s))
pf=pd.read_csv(csv_file_path,sep=',') # 假设数据集以CSV文件形式提供
text_list = pf[['caption_1', 'caption_2', 'caption_3', 'caption_4', 'caption_5']].values.flatten().tolist()
audio_list=pf[['file_name']].values.flatten().tolist()
audio_path=["/data/air/pc/Mobile-Search-Engine/datasets/clotho/evaluation/"+file for file in audio_list]

import argparse

# 创建解析器
parser = argparse.ArgumentParser(description="Your script description")

# 添加命令行参数
#parser.add_argument("audio_num_blocks", type=int, help="Number of audio blocks")

# parser.add_argument("--audio_num_blocks", default=12, type=int, help="Number of audio blocks")
parser.add_argument("--device", type=str, default="cuda:5", help="Device to use (cuda:2 or cpu)")
parser.add_argument("--audio_num_blocks", default=2,type=int, help="Number of audio blocks")
# 解析命令行参数
args = parser.parse_args()

# 获取 audio_num_blocks 的值
audio_num_blocks=args.audio_num_blocks
device = args.device

audio_num_blocks_1=audio_num_blocks
audio_num_blocks_2=12
device_ids = [1,2,3,4,5] 
device = "cuda:6" if torch.cuda.is_available() else "cpu"

#device = "cuda:0" if torch.cuda.is_available() else "cpu"

model_1 = imagebind_model.imagebind_huge(pretrained=True,audio_num_blocks=audio_num_blocks)
model_2 = imagebind_model.imagebind_huge(pretrained=True,audio_num_blocks=audio_num_blocks_2)
v_block=len(model_1.modality_trunks["vision"].blocks)
t_block=len(model_1.modality_trunks["text"].blocks)
a_block=len(model_1.modality_trunks["audio"].blocks)
i_block=len(model_1.modality_trunks["imu"].blocks)
model_1=model_1.cuda()
model_1 = model_1.to(device_ids[0]) 
model_1 = DataParallel(model_1,device_ids=device_ids)
model_1.eval()

model_2=model_2.cuda()
model_2 = model_2.to(device_ids[0]) 
model_2 = DataParallel(model_2,device_ids=device_ids)
model_2.eval()
# model_2.eval()
# model_2.to(device)
import pandas as pd
def run_inference():
    Clotho_dataset = ClothoTextDataset(csv_file=csv_file_path,device=device)
    batch_size=128
    test_dl = DataLoader(dataset=Clotho_dataset, batch_size=batch_size, shuffle=False, drop_last=False,
            num_workers=4, pin_memory=True, persistent_workers=True)
    counts_r1=np.array([])
    counts_r10=np.array([])
    count_ones_r10=0
    batches=[audio_path[i:i+batch_size] for i in range(0,len(audio_path),batch_size)]
    batch=audio_path[0:29]
    audio_embeddings=torch.Tensor().to(device)
    topk1=[1,5, 10, 20, 30, 40, 50, 60,70,80,90,100,110,120,130,300,400,500,600,700,800,900,1000]
    counts_rs = {}
    shortlist={}
    shortlist_item={}
    for k in topk1:
                counts_rs[f'counts_r{k}'] = np.array([])
                shortlist[f'counts_r{k}']=[]
                shortlist_item[f'counts_r{k}']=[]
    embeddings={}
    # with torch.no_grad():
    #         # model_device = next(model_1.module.parameters()).device
    #         # input={  
    #         # ModalityType.AUDIO: data.load_and_transform_audio_data(audio_path,device=model_device)
    #         # }
    #         # embeddings[ModalityType.AUDIO] = model_1(input)[ModalityType.AUDIO]
    #         # 假设你想要保存模型参数和张量
    #         # torch.save({
    #         #     'audio_embeddings': embeddings[ModalityType.AUDIO]
    #         # }, f'embeddings_{a_block}.pth')
    #         # # # 加载保存的内容
    #         checkpoint = torch.load(f'embeddings_{a_block}.pth')
    #         # 获取模型参数和张量
    #         embeddings[ModalityType.AUDIO]= checkpoint['audio_embeddings']


    # with torch.no_grad():
        
    #     for batch_idx, (x, target) in enumerate(test_dl):
    #         target = target.to(device)
    #         inputs = {
    #             ModalityType.TEXT: data.load_and_transform_text(x, device),
    #            # ModalityType.AUDIO: data.load_and_transform_audio_data(audio_path,device=audio_embeddings.device)
    #         }

    #         embeddings_text = model_1(inputs)
    #         #match_value_1 = embeddings[ModalityType.TEXT].to(audio_embeddings.device)@audio_embeddings.T 
    #         #match_value_1 = embeddings[ModalityType.TEXT] @ embeddings[ModalityType.AUDIO].T 
    #         embeddings[ModalityType.AUDIO]=embeddings[ModalityType.AUDIO].to(embeddings_text[ModalityType.TEXT].device)
    #         match_value_1 = embeddings_text[ModalityType.TEXT] @ embeddings[ModalityType.AUDIO].T 
    #         result_1 = torch.softmax(match_value_1, dim=-1)
    #         _, predicted = torch.max(result_1, dim=-1)
    #         _, topk_indices = torch.topk(result_1, k=10, dim=-1)
    #         counts_r1 = np.concatenate([counts_r1, [int(predicted[i] == target[i].to(predicted.device)) for i in range(len(predicted))]])
    #         # #counts_r1 = np.concatenate([counts_r1, [any(predicted[i] == target[i]) for i in range(len(predicted))]])
    #         # #topk_indices=topk_indices.T
    #         counts_r10=np.concatenate([counts_r10, [int(any(topk_indices[i] == target[i].to(predicted.device))) for i in range(len(target))]])
            
    #         top_indices_list = [torch.topk(result_1, k=k, dim=-1)[1] for k in topk1]
            
    #         for k, top_indices, counts_r in zip(topk1, top_indices_list, counts_rs):
    #             if k == 1:
    #                 counts_rs[counts_r] = np.concatenate([counts_rs[counts_r], [int(predicted[i] == target[i].to(predicted.device)) for i in range(len(predicted))]])
    #                 for item in top_indices:
    #                     shortlist_item[counts_r].append([item.item()])
    #             else:
    #                 counts_rs[counts_r] = np.concatenate([counts_rs[counts_r], [int(any(top_indices[i] == target[i].to(predicted.device))) for i in range(len(target))]])
    #                 for i,row in enumerate(top_indices):
    #                     list=[]
    #                     list_item=[]
    #                     for item in row:
    #                         list.append(audio_path[item])
    #                         list_item.append(item.item())
    #                     shortlist[counts_r].append(list)
    #                     shortlist_item[counts_r].append(list_item)
          
    #         # for i,row in enumerate(topk_indices):
    #         #     list=[]
    #         #     list_item=[]
    #         #     for item in row:
    #         #         list.append(audio_path[item])
    #         #         list_item.append(item.item())
    #         #     shortlist.append(list)
    #         #     shortlist_item.append(list_item)
            
    #         r1=(np.sum(counts_rs['counts_r1']==1))/len(counts_rs['counts_r1'])
    #         r5=(np.sum(counts_rs['counts_r5']==1))/len(counts_rs['counts_r1']) 
    #         r10=(np.sum(counts_rs['counts_r10']==1))/len(counts_rs['counts_r1']) 
            
    #         logging.info(f"batch_idx = {batch_idx}, r1={r1},r10={r5}, test_total = {len(counts_rs['counts_r1'])}")
        # np.savetxt(f'./results/clotho/R10/t{t_block}_a{a_block}.txt',counts_r10,fmt='%d')
        # np.savetxt(f'./results/clotho/R1/t{t_block}_a{a_block}.txt',counts_r1,fmt='%d')
        
    # part2
    with open('short_list.json', 'r') as file:
        shortlist_item = json.load(file)
    test_dl2 = DataLoader(dataset=Clotho_dataset, batch_size=1, shuffle=False, drop_last=False,
        num_workers=4, pin_memory=True, persistent_workers=True)
    topk1=[1,5, 10, 20, 30, 40, 50, 60,70,80,90,100,110,120,130,300,400,500,600,700,800,900,1000]
    counts_rs = {}
    embeddings2={}
    with torch.no_grad():
        checkpoint = torch.load(f'embeddings_{audio_num_blocks_2}.pth')
        # 获取模型参数和张量
        embeddings2[ModalityType.AUDIO]= checkpoint['audio_embeddings']
    for k in topk1:
                counts_rs[f'counts_r{k}'] = np.array([])
    batch_size=1
    for counts_r in counts_rs:    
        for batch_idx, (x, target) in enumerate(test_dl2):
            target = target.to(device)
            inputs = {
                ModalityType.TEXT: data.load_and_transform_text(x, device),
            }

            embeddings_text = model_2(inputs)
            embeddings_audio=torch.Tensor()    
            embeddings_audio=embeddings_audio.to(embeddings2[ModalityType.AUDIO].device)      
            for item in shortlist_item[counts_r][batch_idx]: #batch_size=1
                if embeddings_audio.size(0)==0:
                    embeddings_audio=embeddings2[ModalityType.AUDIO][item].unsqueeze(0)
                else:
                    embeddings_audio=torch.cat((embeddings_audio, embeddings2[ModalityType.AUDIO][item].unsqueeze(0)), dim=0)
                #embeddings_audio.append(embeddings2[ModalityType.AUDIO][item])
            
            embeddings_audio=embeddings_audio.to(embeddings_text[ModalityType.TEXT].device)
            match_value_1 = embeddings_text[ModalityType.TEXT] @ embeddings_audio.T 
            result_1 = torch.softmax(match_value_1, dim=-1)
            _, predicted = torch.max(result_1, dim=-1)
            predicted=torch.Tensor([shortlist_item[counts_r][batch_idx*batch_size][predicted]])
            #_, topk_indices = torch.topk(result_1, k=10, dim=-1)
            top_indices_list = [torch.topk(result_1, k=k, dim=-1)[1] if k <= len(shortlist_item[counts_r][batch_idx]) else None for k in topk1]
            for k, top_indices, counts_rk in zip(topk1, top_indices_list, counts_rs):
                if k == 1:
                        counts_rs[counts_rk] = np.concatenate([counts_rs[counts_rk], [int(predicted[i] == target[i].to(predicted.device)) for i in range(len(predicted))]])
                elif k<=len(shortlist_item[counts_r][batch_idx]):            
                        results_k=[]
                        for i in range(k):
                            results_k.append(shortlist_item[counts_r][batch_idx][top_indices[0][i].item()])
                        counts_rs[counts_rk] = np.concatenate([counts_rs[counts_rk], [int(any(results_k[i] == target[i].to(predicted.device))) for i in range(len(target))]])
                        

            r1=(np.sum(counts_r1==1))/len(counts_r1)
            r10=(np.sum(counts_r10==1))/len(counts_r10) 
        
            logging.info(f"batch_idx = {batch_idx}, r1={r1},r10={r10}, test_total = {len(counts_r1)}")
    
        results=[]
        lists=[]
        for counts in counts_rs:
            correct=np.sum(counts_rs[counts] == 1)/len(counts_rs[counts])
            results.append(str(correct))
            lists.append(counts)
        # a_block=len(model_1.modality_trunks["audio"].blocks)
       
        r1=[str(a_block)]
        list1=['audio层数']
        list=list1+lists
        results=r1+results
        import openpyxl
        from openpyxl import Workbook
        # 数据
        data1 = [
         
            results
        ]
        
        # 打开Excel文件
        #workbook = openpyxl.load_workbook('topk-clotho-search.xlsx')
        workbook=Workbook()
        # 选择或创建工作表
        sheet_name = 'Sheet0'  # 请根据实际情况修改工作表名称
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

        # 确定插入的起始行
        start_row = sheet.max_row + 1
        
        # 将数据插入Excel
        for row in data1:
            sheet.append(row)

        # 保存修改后的Excel文件
        workbook.save('topk-clotho-search.xlsx')
    #logging.info(f"batch_idx = {batch_idx}, test_correct = {test_correct}, test_total = {test_total}, Accuracy = {acc}, Recall = {recall}")
    #print(len(counts_r10))
    # count_ones_r1 = np.sum(counts_r1 == 1)
    # count_ones_r10 = np.sum(counts_r10 == 1)
    # r1=count_ones_r1/len(counts_r1)
    # r10=count_ones_r10/len(counts_r1)
    # np.savetxt(f'./results/clotho_t2a/R1/t{t_block}_a{a_block}_acc{r1}.txt',counts_r1,fmt='%d')
    # np.savetxt(f'./results/clotho_t2a/R10/t{t_block}_a{a_block}_acc{r10}.txt',counts_r10,fmt='%d')
    
    return r1,r10

def main():
    Accuracy = run_inference()
    print("Model Performance:", Accuracy)

def print_text_label():
    data_transform = transforms.Compose(
        [
            transforms.Resize(
                224, interpolation=transforms.InterpolationMode.BICUBIC
            ),
            transforms.CenterCrop(224),
            transforms.ToTensor(),
            transforms.Normalize(
                mean=(0.48145466, 0.4578275, 0.40821073),
                std=(0.26862954, 0.26130258, 0.27577711),
            ),
        ]
    )

    datadir = "./.datasets/imagenet"
    test_ds = ImageNet(datadir, split="val", transform=data_transform)
    test_dl = DataLoader(dataset=test_ds, batch_size=64, shuffle=False, drop_last=False,
        num_workers=4, pin_memory=True, persistent_workers=True)
    
    labels = sorted(list(set(batch[1] for batch in test_dl)))
    print(labels)

if __name__ == "__main__":
    main()
    # print_text_label()
