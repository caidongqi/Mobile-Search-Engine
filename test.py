import logging
import torch
import data
import torchvision
import torchmetrics
import time
import csv

from models import imagebind_model
from models.imagebind_model import ModalityType, load_module
from models import lora as LoRA

from pycocotools.coco import COCO
from torchvision import transforms
from torch.utils.data import DataLoader
from torchvision.datasets import CocoDetection

from torchvision import transforms
from torchvision.datasets import ImageNet
from torch.utils.data import DataLoader
from datasets.imagenet import ImageNetDataset
from datasets.coco import CoCoDataset
from metrics.accuracy import Accuracy
import openpyxl
logging.basicConfig(level=logging.INFO, force=True)


lora = False
linear_probing = False
device="cuda:5"
# device = "cpu" if torch.cuda.is_available() else "cpu"
load_head_post_proc_finetuned = True
imagenet_datadir = "/data/yx/ImageBind/.datasets/imagenet"
coco_datadir="/data/air/pc/ImageBind/dataset/val2017/val2017"
test_coco="/data/air/pc/ImageBind/dataset/tempoimage"
datadir1='/data/air/pc/Mobile-Search-Engine/datasets/imagenet-10'
datadir2="/data/air/pc/i-Code/i-Code-V3/dataset/tempo-10"
coco_annotation_file='/data/air/pc/ImageBind/dataset/annotations_trainval2017/annotations/instances_val2017.json'
lora_dir = ''


# workbook = openpyxl.Workbook()

# # 创建一个工作表
# worksheet = workbook.active

# # 保存工作簿为xlsx文件
# workbook.save('example.xlsx')
# data = [
#         ['vision层数', 'text层数', 'audio层数', 'imu层数', 'Accuracy','创建model时间', '合并input时间', '单独输入vision时间', '单独输入text时间', ],
#     #     [ str(v_block), str(t_block), str(a_block), str(i_block), str(total_acc),str(initial_model_time), str(model_merge_time), str(vision_time), str(text_time), str(match_time),
#     #    ]
#     ]

#     # 打开Excel文件



assert not (linear_probing and lora), \
            "Linear probing is a subset of LoRA training procedure for ImageBind. " \
            "Cannot set both linear_probing=True and lora=True. "

if lora and not load_head_post_proc_finetuned:
    # Hack: adjust lora_factor to the `max batch size used during training / temperature` to compensate missing norm
    lora_factor = 12 / 0.07
else:
    # This assumes proper loading of all params but results in shift from original dist in case of LoRA
    lora_factor = 1

initial_model_v_start=time.time()
# Instantiate model
model_v = imagebind_model.imagebind_huge(pretrained=True,vision_embed_dim=1280,vision_num_blocks=32,
                   vision_num_heads=16,text_embed_dim=1024,text_num_blocks=0,
                   text_num_heads=16,out_embed_dim=1024,audio_drop_path=0.1,imu_drop_path=0.7)
initial_model_v_end=time.time()
initial_model_v_time=initial_model_v_end-initial_model_v_start
v_block=len(model_v.modality_trunks["vision"].blocks)
t_block=len(model_v.modality_trunks["text"].blocks)
a_block=len(model_v.modality_trunks["audio"].blocks)
i_block=len(model_v.modality_trunks["imu"].blocks)
if lora:
    model_v.modality_trunks.update(
        LoRA.apply_lora_modality_trunks(model_v.modality_trunks, rank=4,
                                        layer_idxs={ModalityType.TEXT: [ 1, 2, 3, 4, 5, 6, 7, 8],
                                                    ModalityType.VISION: [1, 2, 3, 4, 5, 6, 7, 8]},
                                        modality_names=[ModalityType.TEXT, ModalityType.VISION]))

    # Load LoRA params if found
    LoRA.load_lora_modality_trunks(model_v.modality_trunks,
                                   checkpoint_dir=lora_dir)

    if load_head_post_proc_finetuned:
        # Load postprocessors & heads
        load_module(model_v.modality_postprocessors, module_name="postprocessors",
                    checkpoint_dir=lora_dir)
        load_module(model_v.modality_heads, module_name="heads",
                    checkpoint_dir=lora_dir)
elif linear_probing:
    # Load heads
    load_module(model_v.modality_heads, module_name="heads",
                checkpoint_dir=lora_dir)

model_v.eval()
model_v.to(device)



initial_model_t_start=time.time()
# Instantiate model
model_t = imagebind_model.imagebind_huge(pretrained=True,vision_embed_dim=1280,vision_num_blocks=0,
                   vision_num_heads=16,text_embed_dim=1024,text_num_blocks=24,
                   text_num_heads=16,out_embed_dim=1024,audio_drop_path=0.1,imu_drop_path=0.7)
initial_model_t_end=time.time()
initial_model_t_time=initial_model_t_end-initial_model_t_start

if lora:
    model_t.modality_trunks.update(
        LoRA.apply_lora_modality_trunks(model_t.modality_trunks, rank=4,
                                        layer_idxs={ModalityType.TEXT: [ 1, 2, 3, 4, 5, 6, 7, 8],
                                                    ModalityType.VISION: [1, 2, 3, 4, 5, 6, 7, 8]},
                                        modality_names=[ModalityType.TEXT, ModalityType.VISION]))

    # Load LoRA params if found
    LoRA.load_lora_modality_trunks(model_t.modality_trunks,
                                   checkpoint_dir=lora_dir)

    if load_head_post_proc_finetuned:
        # Load postprocessors & heads
        load_module(model_t.modality_postprocessors, module_name="postprocessors",
                    checkpoint_dir=lora_dir)
        load_module(model_t.modality_heads, module_name="heads",
                    checkpoint_dir=lora_dir)
elif linear_probing:
    # Load heads
    load_module(model_t.modality_heads, module_name="heads",
                checkpoint_dir=lora_dir)

model_t.eval()
model_t.to(device)


def run_inference():
    data_transform = transforms.Compose(
        [
            transforms.Resize(
                224, interpolation=transforms.InterpolationMode.BICUBIC
            ),
            transforms.CenterCrop(224),
            transforms.ToTensor(),
            transforms.Normalize(
                mean=(0.48145466, 0.4578275, 0.40821073),
                std=(0.26862954, 0.26130258, 0.27577711),
            ),
        ]
    )

    test_ds1=CoCoDataset(datadir=coco_datadir, annFile=coco_annotation_file,transform=data_transform)
    #test_ds = ImageNetDataset(datadir=test_coco, split="val", transform=data_transform)
    test_dl1=DataLoader(dataset=test_ds1, batch_size=64, shuffle=False, drop_last=False, num_workers=4, pin_memory=True, persistent_workers=True)
    #test_dl = DataLoader(dataset=test_ds, batch_size=64, shuffle=False, drop_last=False, num_workers=4, pin_memory=True, persistent_workers=True)
    import os
    test_acc = Accuracy(task="multiclass", num_classes=1000, average="micro",device=device)
    file_count = sum(1 for entry in os.scandir(coco_datadir) if entry.is_file())
    test_correct = 0
    test_total = 0
    import data
    model_merge_executed = False
    batch_executed=False
    total_correct=0
    batch_time=0
    print(len(test_ds1))
    with torch.no_grad():
        for batch_idx, (x, target) in enumerate(test_dl1):
            if not batch_executed:
                batch_start=time.time()
            x = x.to(device)
            target=[t.to(device) for t in target]
            
            inputs = {
                ModalityType.VISION: x,
                ModalityType.TEXT: data.load_and_transform_text(test_ds1.text_list, device),
            }
            vision_inputs={
                ModalityType.VISION: x
            }
            text_inputs={ ModalityType.TEXT: data.load_and_transform_text(test_ds1.text_list, device),
            }
            
            # embeddings_v=model_v(vision_inputs)
            # embeddings_t=model_t(text_inputs)
            
            #分开算 合并算  
            if  not model_merge_executed:
                    # model_merge_start=time.time()
                    # embeddings = model(inputs)
                    # model_merge_end=time.time()
                    # model_merge_time=model_merge_end-model_merge_start
                    
                    vision_start=time.time()
                    vision_embeddings=model_v(vision_inputs)
                    vision_end=time.time()
                    vision_time=vision_end-vision_start
                    
                    text_start=time.time()
                    text_embeddings=model_t(text_inputs)
                    text_end=time.time()
                    text_time=text_end-text_start
                    
                    match_start=time.time()
                    match_value_1 = vision_embeddings[ModalityType.VISION]@text_embeddings[ModalityType.TEXT].T * (lora_factor if lora else 1)
                    match_end=time.time()
                    match_time=match_end-match_start
                    model_merge_executed=True
            vision_embeddings=model_v(vision_inputs)
            text_embeddings=model_t(text_inputs)        
            match_value_1 = vision_embeddings[ModalityType.VISION]@text_embeddings[ModalityType.TEXT].T * (lora_factor if lora else 1)
            print(match_value_1.shape)
            # num=5
            # topk_indices = torch.topk(match_value_1, num, dim=-1).indices
            # # 获取每行最大值的索引
            # max_indices = torch.argmax(match_value_1, dim=1)
            result_1 = torch.softmax(match_value_1, dim=-1)
            _, predicted = torch.max(result_1, -1)
            _,top_k_indices = torch.topk(result_1, k=5, dim=-1)
            correct=0
            
            for i in range(top_k_indices.size(0)):
                    row_tensor1 = top_k_indices[i, :]
                    row_tensor1= torch.add(row_tensor1, 1)
                    # 提取每个子tensor的第i列
                    column_values = [tensor[i] for tensor in target]
                    extracted_values = [item.item() for item in column_values]

                    # 组成新的tensor
                    new_tensor = torch.cat([torch.tensor(extracted_values)])
                    new_tensor=torch.unique(new_tensor.to(device))
                    intersection = torch.unique(torch.cat((row_tensor1, new_tensor)))
                    #intersection = torch.intersection(row_tensor1, new_tensor)
                    if len(intersection)<len(row_tensor1) + len(new_tensor):
                         correct+=1
            acc=correct/64
            total_correct+=correct
            # acc = test_acc.update(result_1.argmax(1), target)
            # # intersection = set() & set(arr2)
            # # acc=(0 if not set(predicted)&set(target))
            # correct = predicted.eq(target).sum()
            # test_correct += correct.item()
            # test_total += target.size(0)
            logging.info(f"batch_idx = {batch_idx}, test_correct = {test_correct}, test_total = {test_total}, Accuracy = {acc}")
            
            if not batch_executed:
                batch_end=time.time()
                batch_time=batch_end-batch_start
                batch_executed=True
    #total_acc = test_acc.compute()
    
    total_acc=total_correct/file_count
    logging.info(f"Accuracy = {total_acc}")
    
    #将时间记录下来
    v_block=len(model_v.modality_trunks["vision"].blocks)
    t_block=len(model_t.modality_trunks["text"].blocks)
    a_block=len(model_v.modality_trunks["audio"].blocks)
    i_block=len(model_v.modality_trunks["imu"].blocks)
    filename=f'v{v_block}_t{t_block}.xlsx'
    
    import openpyxl

    # 数据
    data = [
        ['vision层数', 'text层数', 'audio层数' ,'imu层数','Accuracy','创建model_v时间','创建model_t时间',  '单独输入vision时间', '单独输入text时间','match时间' ,'一个batch','device'],
        [ str(v_block), str(t_block),str(a_block), str(i_block),str(total_acc),str(initial_model_v_time),str(initial_model_t_time) , str(vision_time), str(text_time), str(match_time),str(batch_time),str(device)
       ]
    ]

    # 打开Excel文件
    workbook = openpyxl.load_workbook('example.xlsx')

    # 选择或创建工作表
    sheet_name = 'Sheet1'  # 请根据实际情况修改工作表名称
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

    # 确定插入的起始行
    start_row = sheet.max_row + 1

    # 将数据插入Excel
    for row in data:
        sheet.append(row)

    # 保存修改后的Excel文件
    workbook.save('example.xlsx')

    
    # # 写入CSV文件
    # with open(filename, 'w', newline='') as file:
    #     writer = csv.writer(file)
    #     writer.writerows(data)

    # print(f"数据已写入CSV文件: {filename}")

   
    
    


if __name__ == "__main__":
    run_inference()
